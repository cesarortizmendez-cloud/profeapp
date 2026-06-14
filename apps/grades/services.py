"""
Servicios de cálculo de notas para ProfeApp.
Maneja promedios ponderados, estadísticas y exportación.
"""
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Avg, Count, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from .models import Grade, GradeColumn, GradeReport


def calculate_weighted_average(student, course):
    """
    Calcula el promedio ponderado de un alumno en un curso.
    
    - Solo considera columnas publicadas.
    - Si la suma de pesos publicados < 100, calcula sobre los pesos disponibles.
    - Retorna (promedio: Decimal | None, detalle: list)
    """
    columns = GradeColumn.objects.filter(
        course=course,
        is_published=True,
        weight__gt=0,
    ).order_by('order', 'created_at')

    if not columns.exists():
        return None, []

    total_weight = Decimal('0')
    weighted_sum = Decimal('0')
    detail = []

    for col in columns:
        grade = Grade.objects.filter(student=student, column=col).first()
        score = None
        is_absent = False

        if grade:
            score = grade.get_effective_score()
            is_absent = grade.is_absent
        
        weight = col.weight  # Peso en %
        detail.append({
            'column': col,
            'score': score,
            'weight': weight,
            'is_absent': is_absent,
        })

        if score is not None:
            total_weight += weight
            weighted_sum += score * weight

    if total_weight == 0:
        return None, detail

    average = (weighted_sum / total_weight).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
    return average, detail


def update_grade_report(student, course):
    """Actualiza (o crea) el GradeReport de un alumno en un curso."""
    average, _ = calculate_weighted_average(student, course)
    
    report, _ = GradeReport.objects.update_or_create(
        student=student,
        course=course,
        defaults={
            'weighted_average': average,
            'is_passing': (average >= Decimal('4.0')) if average is not None else None,
        }
    )
    return report


def recalculate_course_reports(course):
    """Recalcula los reportes de todos los alumnos de un curso."""
    from apps.courses.models import Enrollment
    enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student')
    for enrollment in enrollments:
        update_grade_report(enrollment.student, course)


def get_course_statistics(course):
    """Estadísticas generales del curso."""
    from apps.courses.models import Enrollment
    
    students_count = Enrollment.objects.filter(course=course, is_active=True).count()
    reports = GradeReport.objects.filter(course=course).exclude(weighted_average__isnull=True)
    
    if not reports.exists():
        return {
            'students_count': students_count,
            'graded_count': 0,
            'average': None,
            'passing_count': 0,
            'failing_count': 0,
            'passing_rate': 0,
        }

    avg_data = reports.aggregate(avg=Avg('weighted_average'))
    passing = reports.filter(is_passing=True).count()
    failing = reports.filter(is_passing=False).count()
    graded = reports.count()

    return {
        'students_count': students_count,
        'graded_count': graded,
        'average': avg_data['avg'],
        'passing_count': passing,
        'failing_count': failing,
        'passing_rate': round((passing / graded * 100), 1) if graded > 0 else 0,
    }


def export_grades_to_excel(course):
    """
    Genera un archivo Excel con las notas del curso.
    Retorna un BytesIO con el archivo.
    """
    from apps.courses.models import Enrollment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Notas - {course.name[:28]}'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    subheader_fill = PatternFill(start_color='E8ECFD', end_color='E8ECFD', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

    # Título del reporte
    ws.merge_cells('A1:A1')
    ws['A1'] = f'REPORTE DE NOTAS - {course.name.upper()}'
    ws['A1'].font = Font(bold=True, size=13)
    ws.row_dimensions[1].height = 25

    columns = list(GradeColumn.objects.filter(course=course).order_by('order', 'created_at'))
    
    # Fila de encabezados
    headers = ['N°', 'RUT', 'Apellidos', 'Nombre']
    for col in columns:
        headers.append(col.name)
    headers.extend(['Promedio Ponderado', 'Estado'])
    
    # Fila de pesos
    weights_row = ['', '', '', 'Ponderación (%)']
    for col in columns:
        weights_row.append(f'{col.weight}%')
    weights_row.extend(['', ''])

    # Escribir encabezados (fila 2)
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'].alignment = center_align
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Escribir pesos (fila 3)
    for col_idx, weight in enumerate(weights_row, 1):
        cell = ws.cell(row=3, column=col_idx, value=weight)
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = border
        cell.font = Font(bold=True, size=9, italic=True)

    # Datos de alumnos
    enrollments = Enrollment.objects.filter(
        course=course, is_active=True
    ).select_related('student').order_by('student__last_name', 'student__first_name')

    row_num = 4
    for idx, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        average, detail = calculate_weighted_average(student, course)

        row_data = [
            idx,
            student.rut or '',
            student.last_name,
            student.first_name,
        ]

        for grade_detail in detail:
            score = grade_detail['score']
            row_data.append(float(score) if score is not None else '')

        row_data.append(float(average) if average is not None else '')
        row_data.append('APROBADO' if (average and average >= 4) else ('REPROBADO' if average else 'S/N'))

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = center_align if col_idx > 2 else left_align
            cell.border = border

            # Colorear notas
            if col_idx > 4 and isinstance(value, float):
                if value >= 4.0:
                    cell.fill = pass_fill
                else:
                    cell.fill = fail_fill

        # Colorear estado
        state_cell = ws.cell(row=row_num, column=len(headers))
        if state_cell.value == 'APROBADO':
            state_cell.fill = pass_fill
            state_cell.font = Font(bold=True, color='155724')
        elif state_cell.value == 'REPROBADO':
            state_cell.fill = fail_fill
            state_cell.font = Font(bold=True, color='721C24')

        row_num += 1

    # Ajustar anchos de columna
    col_widths = [5, 14, 20, 18] + [12] * len(columns) + [20, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'E4'

    # Hoja de estadísticas
    ws2 = wb.create_sheet('Estadísticas')
    stats = get_course_statistics(course)
    ws2['A1'] = 'ESTADÍSTICAS DEL CURSO'
    ws2['A1'].font = Font(bold=True, size=12)
    stats_data = [
        ('Total de alumnos', stats['students_count']),
        ('Alumnos con notas', stats['graded_count']),
        ('Promedio del curso', float(stats['average']) if stats['average'] else 'Sin datos'),
        ('Aprobados', stats['passing_count']),
        ('Reprobados', stats['failing_count']),
        ('Tasa de aprobación', f"{stats['passing_rate']}%"),
    ]
    for i, (label, value) in enumerate(stats_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_students_from_excel(file, course, professor):
    """
    Importa alumnos desde un Excel.
    Columnas esperadas: RUT, Apellidos, Nombre, Email (opcional)
    Retorna (created, updated, errors)
    """
    from apps.accounts.models import User
    from apps.courses.models import Enrollment
    import pandas as pd

    created = []
    updated = []
    errors = []

    try:
        df = pd.read_excel(file, dtype=str)
        df.columns = df.columns.str.strip().str.lower()

        # Mapeo flexible de columnas
        col_map = {}
        for col in df.columns:
            col_clean = col.strip().lower()
            if any(k in col_clean for k in ['rut', 'run']):
                col_map['rut'] = col
            elif any(k in col_clean for k in ['apellido', 'last']):
                col_map['last_name'] = col
            elif any(k in col_clean for k in ['nombre', 'first', 'name']) and 'apellido' not in col_clean:
                col_map['first_name'] = col
            elif any(k in col_clean for k in ['email', 'correo', 'mail']):
                col_map['email'] = col

        if 'rut' not in col_map:
            errors.append('No se encontró columna RUT/RUN en el archivo.')
            return created, updated, errors

        for idx, row in df.iterrows():
            row_num = idx + 2
            try:
                rut = str(row.get(col_map.get('rut', ''), '')).strip()
                last_name = str(row.get(col_map.get('last_name', ''), '')).strip()
                first_name = str(row.get(col_map.get('first_name', ''), '')).strip()
                email = str(row.get(col_map.get('email', ''), '')).strip() if 'email' in col_map else ''

                if not rut or rut == 'nan':
                    continue

                # Limpiar RUT
                rut = rut.replace('.', '').replace(' ', '').upper()
                
                # Generar username desde RUT
                username = rut.replace('-', '').lower()
                
                # Crear o actualizar usuario
                user, was_created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'role': 'student',
                        'rut': rut,
                        'last_name': last_name,
                        'first_name': first_name,
                        'email': email if email and email != 'nan' else '',
                        'must_change_password': True,
                    }
                )

                if was_created:
                    user.set_password(rut)  # Contraseña inicial = RUT
                    user.save()
                    created.append(user.get_full_name() or username)
                else:
                    # Actualizar datos si ya existe
                    changed = False
                    if last_name and user.last_name != last_name:
                        user.last_name = last_name
                        changed = True
                    if first_name and user.first_name != first_name:
                        user.first_name = first_name
                        changed = True
                    if email and email != 'nan' and user.email != email:
                        user.email = email
                        changed = True
                    if changed:
                        user.save()
                    updated.append(user.get_full_name() or username)

                # Matricular en el curso
                Enrollment.objects.get_or_create(course=course, student=user)

            except Exception as e:
                errors.append(f'Fila {row_num}: {str(e)}')

    except Exception as e:
        errors.append(f'Error al leer el archivo: {str(e)}')

    return created, updated, errors
